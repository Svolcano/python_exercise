import time
from openpyxl import load_workbook, Workbook


def get_name_phone(file_path, sheet_name, name_prefix, phone_refix):
    result = {}
    book = load_workbook(file_path, read_only=True)
    sheet = book[sheet_name]
    max_row = sheet.max_row
    i = 2
    while i < max_row:
        name = sheet["%s%d" % (name_prefix, i)].value
        phone = sheet["%s%d" % (phone_refix, i)].value
        result[name] = phone
        i += 1
    book.close()
    return result



def mix_save(target_file_name, src_file, src_sheet_name, name_phone):
    target_book = Workbook(write_only=True)
    src_file = load_workbook(src_file, read_only=True)
    src_sheet = src_file[src_sheet_name]
    max_row = src_sheet.max_row
    max_column = src_sheet.max_column
    src_all_rows = list(src_sheet.rows)
    dest_sheet = target_book.create_sheet("done", 1)
    for j in range(1, max_column + 1):
        dest_sheet.cell(1, j).value = src_all_rows[0][j].value
    dest_sheet.cell(1, max_column+1).value = '电话号'
    i = 2
    for r in src_all_rows[1:]:
        j = 0
        for c in r:
            if c.col_idx == 2:
                name = c.value
                ph = name_phone.get(name, '-1')
            dest_sheet.cell(i, j).value = c.value
            j += 1
        dest_sheet.cell(i, j+1).value = ph
    target_book.save(target_file_name)

if __name__ == "__main__":
    name_ph = {}
    st = time.time()
    name_ph = get_name_phone('截至4月17日8点会员数据.xlsx', 'Sheet1', 'B', 'E')
    et1 = time.time()
    print("read1 cost : %.3f" % (et1-st))
    mix_save("done.xlsx", '回款日志汇总.xlsx', 'Sheet1', name_ph)
    et2 = time.time()
    print("read2 cost : %.3f" % (et2 - et1))

